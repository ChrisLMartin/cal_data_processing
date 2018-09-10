# -*- coding: utf-8 -*-
"""
Created on Fri Dec 22 16:47:14 2017

@author: Christopher Martin

v3.00

20180910

Files to be imported are selected (can be multiple files) and dragged and 
dropped onto "cal_data_import.bat". Errors and warnings are written to 
"error_cal_data_import.txt".

Previous use required "replace_degree.py" to be run on folder containing 
Calmetrix cc2 export TSV files to be imported (so that degree symbol is 
replaced with "deg" preventing issues with Windows command prompt), then 

TO DO: 
    - search calmetrix export folder for new files that haven't been processed
        instead of drag and drop on batch file
    - resolve problems with sample_id caused by underscore in sample ID
    - resolve problems with df_param and df_val caused by newlines in CC2 
        logger details fields
    - check if any efc_calcs() values are/should be zero
    - refactor/combine write_to_excel() and append_df_to_excel()
    - refactor (potentially with dict of dfs for list of imported files in 
        main()) in order to reduce excel read/write occurences/times
    - try/except for initial data import? 
    - move to SQL database instead of excel (SQLite?)
"""

import argparse
from datetime import datetime
import numpy as np
import os
import pandas as pd
import sys


# Used with .bat file for drag and drop
parser = argparse.ArgumentParser()
parser.add_argument('file', nargs='*')
args = parser.parse_args()
files = list(vars(args).values())[0]

'''
DEBUG FILE
'''
#files = [os.path.normpath(
#        'C:/Users/christopher.martin/Documents/Python/cal_data/2018-09/'
#        + '20180904-01_23degC_Ch01_2018-09-04_10-14-37.txt')]


def main(file_list):
    """
    Takes a [file_list] of Calmetrix cc2 export TSV files and calls data import 
    functions on each file, with results saved to excel workbook.
    
    Parameters:
        file_list: List of Calmetrix cc2 export TSV files, with filenames in 
            the form "/[sample-id]_[xx]degC_Ch[x]_[yyy-mm-dd]_[hh-mm-ss].txt"
            DO NOT USE UNDERSCORES IN SAMPLE-ID
    Returns: None
    """
    for file in file_list:
        sample_id, sample_type, out_filename, out_location = check_name(file)
        df_p, df_v, df_vp = data_in(file, sample_id)
        if sample_type == "EFC":
            m_bind = efc_calcs(df_p)
        elif sample_type == "OPC":
            m_bind = opc_calcs(df_p)
        df_v = tidy_val_df(df_v, m_bind)
        df_p = tidy_param_df(sample_id, df_p, out_filename)
        write_to_excel(sample_id, df_p, df_v, df_vp, out_location)
    
def check_name(input_filename):
    """
    Parameters:
        input_filename: Calmetrix cc2 export TSV file, with filename in the 
        form "/[sample-id]_[xx]degC_Ch[x]_[yyy-mm-dd]_[hh-mm-ss].txt"
        DO NOT USE UNDERSCORES IN SAMPLE-ID
    Returns:
        sample_id: id from [input_filename], all characters before first 
            underscore
        sample_type: EFC or OPC
        output_excel_location: file location of excel output file
    """
     # Underscores in sample ID cause problems here
    # Check to make sure using geopolymer sample naming system
    sample_id = os.path.basename(input_filename).split('_')[0]
    year = datetime.today().year
    short_year = str(year)[-2:]
    if sample_id.startswith(str(year)):
        sample_type = "EFC"
    elif sample_id.startswith(short_year):
        sample_type = "OPC"
    else:
        print('Unexpected sample id (does not start with {} or {}): {}'.format(year, short_year, sample_id))
        
        valid = {"efc": "EFC", "e": "EFC", "ef": "EFC",
                 "opc": "OPC", "op": "OPC", "o": "OPC"}
        question = "Please enter a sample type."
        prompt = " [efc/opc] "
    
        while True:
            sys.stdout.write(question + prompt)
            choice = input().lower()
            if choice in valid:
                sample_type = valid[choice]
                break
            else:
                sys.stdout.write("Please respond with 'efc' or 'opc' "
                                 "(or 'e' or 'o').\n")
                
    output_excel_filename = {"EFC": "CalorimetryData2018Automated.xlsx", 
                             "OPC": "CalorimetryDataOPCAutomated.xlsx"}
    out_filename = output_excel_filename[sample_type]
    output_excel_location = os.path.normpath(
            'C:/Users/christopher.martin/Documents/Python/cal_data_processing/{}'.format(out_filename))
#        sys.exit()
    
    return sample_id, sample_type, out_filename, output_excel_location


def data_in(input_filename, sample_id):
    """
    Reads data from [input_filename] of type Calmetrix cc2 export TSV into
    pandas DataFrames.
    
    Parameters:
        input_filename: Calmetrix cc2 export TSV file, with filename in the 
           form "/[sample-id]_[xx]degC_Ch[x]_[yyy-mm-dd]_[hh-mm-ss].txt"
           DO NOT USE UNDERSCORES IN SAMPLE-ID
        sample_id: id from [input_filename], all characters before first 
            underscore
    Returns:
        df_param_indexed: long df of mix parameters, with parameter names as 
            index. If newlines have been used in any field of CC2 logger 
            details fields, there will be problems reading in to df_param and 
            df_val
        df_val: df of all recorded time, power, and energy values
        df_val_params: df with sample id for first rows of excel sheets, to 
            allow simpler excel graphing and annotation
    """
   
    print('Processing sample {}'.format(sample_id))
    # Encoding set to latin1 due to presence of degree symbol
    # newlines in CC2 logger details fields will cause issues
    header_rows = 30
    df_param = pd.read_table(input_filename, 
                             nrows=header_rows, 
                             encoding="latin1", 
                             header=None)
    df_val = pd.read_table(input_filename, skiprows=header_rows)
    df_param_indexed = df_param.set_index(0)
#    redundant due to parsing sample id from filename
#    sample_id = df_param_indexed.loc['Sample ID', 1]
    
    d1 = {1: pd.Series(['', ''], index=['Sample ID', 'Label'])}
    df_val_params = pd.DataFrame(d1)
    df_val_params.loc['Sample ID', 1] = sample_id
    df_val_params.loc['Label', 1] = sample_id

    return df_param_indexed, df_val, df_val_params


def efc_calcs(df_param_indexed):
    """
    Takes [sample_id] and copies of [df_param_indexed] and [df_val] from 
    data_in() and removes data prior to isothermal (first local minimum), 
    calculates specific power and energy values, and adds a hyperlink to the
    excel parameter row for easier navigation between sheets.
    
    Parameters:
        sample_id: from data_in(), id of calorimetry sample
        df_param_indexed: df of mix parameters, with parameter names as index
        df_val: df of all recorded time, power, and energy values
    Returns:
        df_param_indexed_transpose: wide df of parameters, with excel hyperlink
            for worksheet navigation
        df_val: df of calorimetry values, starting from isothermal (if found)
            with energy and power values per mass of binder
    """
    
    df_param_indexed = df_param_indexed.copy()
    
    ''' commented 20180210 after Calmetrix update
    # Remove for cc1 data exported with cc2
    mix_start = datetime.strptime(
            df_param_indexed.loc['Mix Time', 1], "%d-%b-%Y %H:%M:%S")
    log_start = datetime.strptime(
            df_param_indexed.loc['Start Time', 1], "%d-%b-%Y %H:%M:%S")
    time_difference = (log_start - mix_start).total_seconds()
    '''

    # Calculate mass of binder in sample
    m_slag = float(df_param_indexed.loc['Suppl 1 Mass, g', 1])
    m_fa = float(df_param_indexed.loc['Suppl 2 Mass, g', 1])
    m_water = float(df_param_indexed.loc['Water Mass, g', 1])
    m_agg = float(df_param_indexed.loc['Aggr Mass, g', 1])
    m_sample = float(df_param_indexed.loc['Sample Mass, g', 1])
    m_sample_scm = m_sample / (m_slag + m_fa + m_water + m_agg) * (m_slag + m_fa)
    
    return m_sample_scm


def opc_calcs(df_param_indexed):
    '''
    Takes [sample_id] and copies of [df_param_indexed] and [df_val] from 
    data_in() and removes data prior to isothermal (first local minimum), 
    calculates specific power and energy values, and adds a hyperlink to the
    excel parameter row for easier navigation between sheets.
    
    Parameters:
        sample_id: from data_in(), id of calorimetry sample
        df_param_indexed: df of mix parameters, with parameter names as index
        df_val: df of all recorded time, power, and energy values
    Returns:
        df_param_indexed_transpose: wide df of parameters, with excel hyperlink
            for worksheet navigation
        df_val: df of calorimetry values, starting from isothermal (if found)
            with energy and power values per mass of binder
    '''
    
    df_param_indexed = df_param_indexed.copy()
    
    '''  commented 20180210 after Calmetrix update
    # Remove for cc1 data exported with cc2
    mix_start = datetime.strptime(
            df_param_indexed.loc['Mix Time', 1], "%d-%b-%Y %H:%M:%S")
    log_start = datetime.strptime(
            df_param_indexed.loc['Start Time', 1], "%d-%b-%Y %H:%M:%S")
    time_difference = (log_start - mix_start).total_seconds()
    '''
    # select values from DataFrame and calculate mass of binder in sample
    # may be worth checking if any of these values are 0 at some point in the future
    
    m_water = float(df_param_indexed.loc['Water Mass, g', 1])
    m_cem = float(df_param_indexed.loc['Cement Mass, g', 1])
    m_sample = float(df_param_indexed.loc['Sample Mass, g', 1])
    m_sample_cem = m_sample / (m_cem + m_water) * m_cem
    
    return m_sample_cem

    
def tidy_val_df(df_val, m_sample_scm):
    
    df_val = df_val.copy()
    # Set values to 0 prior to isothermal
    # look from min_search_start minutes to min_search_end minutes
    min_search_start = 60 
    min_search_end = 600
    idx_min = np.argmin(
            df_val['Power1,W'].values[min_search_start:min_search_end]) + min_search_start
    if idx_min >= 599:
        idx_min = 0
#    idx_min = 0
    df_val = df_val[idx_min:]
    df_val['Heat1,J'] = df_val['Heat1,J'].apply(lambda x: x - df_val['Heat1,J'].values[0])
    
    # create time in decimal days for RG charts 20180111
    # header names require numbers for cc1 data exported with cc2
    df_val['Power/SCM,W/g'] = df_val['Power1,W'].values / m_sample_scm
    df_val['Heat/SCM,J/g'] = df_val['Heat1,J'].values / m_sample_scm
#    df_val['Tmix,s'] = df_val['Tlog,s'].values + time_difference
    df_val['Tmix,days'] = df_val['Tmix1,s'].values / 86400  # 60 * 60 * 24
#    df_val = df_val.drop('Tlog,s', axis=1)  # remove for cc1 data exported with cc2

#    rearrange columns to place Tmixs first 
    cols = df_val.columns.tolist()
    cols = cols[0:1] + cols[-1:] + cols[1:-1]
#    cols = cols[0:1] + cols[-1:] + cols[1:-1] # For cc1 data exported with cc2
    df_val = df_val[cols]
    
    return df_val

def tidy_param_df(sample_id, df_param_indexed, out_filename):
    df_param_indexed = df_param_indexed.copy()
#    add link to each sheet in excel on paramters sheet, goes to label cell B2 20180111
    d2 = {1 : pd.Series(
            '=HYPERLINK("[{}]\'{}\'!B2", "Sheet")'.format(
                    out_filename, sample_id), index=['Link'])}
    df_param_link = pd.DataFrame(d2)
    df = df_param_link.append(
            df_param_indexed).transpose()
    df = df.drop(columns='Sample Number')

    return df


def write_to_excel(sample_id, 
                   df_param_indexed_transpose, 
                   df_val, 
                   df_val_params, 
                   output_filename):
    """
    Takes [sample_id] from data_in(), [df_param_indexed_transpose], [df_val], 
    and [df_val_params] from efc_calcs(), and [output_filename] from main().
    Writes [df_param_indexed_transpose] to [param_sheet], and [df_val] and 
    [df_val_params] to spreadsheet titled [sample_id] in [output_filename].
    Checks if workbook, sheets, and entries exist prior to writing, prompts to
    replace.
    
    Parameters:
        sample_id: from data_in(), id of calorimetry sample
        df_param_indexed_transpose: wide df of mix parameters
        df_val: df of calorimetry values
        df_val_params: df with sample id for first rows of excel sheets
        output_filename:  excel workbook "*.xlsx". 
            Default: output_excel_location
    Returns: None
    """
    param_sheet = 'Parameters'

    # try to open existing workbook: File not found -> create, add parameter sheet with header, add value sheet
    # if file is found, check if current sample exists; if not, write parameters without header, add value   
    try:
        print('Trying to open workbook {}.'.format(output_filename))
        df_existing_param = pd.read_excel(output_filename, 
                                          sheet_name=param_sheet)
        if sample_id not in df_existing_param.values:
            print('Writing parameters {} ...'.format(sample_id), end=' ', flush=True)
            start_action = datetime.now()
            append_df_to_excel(output_filename, 
                               df_param_indexed_transpose, 
                               param_sheet, 
                               index=False, 
                               header=False)
            stop_action = datetime.now()
            duration = stop_action - start_action
            print('Duration: {} seconds.'.format(duration.total_seconds()))
        else:
            # should add check that there aren't already multiple paramter rows with same id, which can
            # occur if the filename and sample id in file don't match (filename manually changed)
            idx_sample_id = df_existing_param.index[df_existing_param[
                    'Sample ID'] == sample_id][0]
            print(
                'A parameter row with Sample ID {} already exists in this workbook.'.format(sample_id))
            overwrite_parameter = query_yes_no(
                    'Do you want to overwrite this parameter row?')
            if overwrite_parameter:
                print('Overwriting parameters {} ...'.format(sample_id), end=' ', flush=True)
                start_action = datetime.now()
                append_df_to_excel(output_filename, 
                                   df_param_indexed_transpose, 
                                   param_sheet, 
                                   startrow=idx_sample_id+1, 
                                   index=False, 
                                   header=False)
                stop_action = datetime.now()
                duration = stop_action - start_action
                print('Duration: {} seconds.'.format(duration.total_seconds()))
    except FileNotFoundError:
        print('Creating workbook {}'.format(output_filename))
        append_df_to_excel(output_filename, 
                           df_param_indexed_transpose, 
                           param_sheet, 
                           index=False)
        pass

    wb = pd.ExcelFile(output_filename)

    if sample_id not in wb.sheet_names:
        print('Writing values sheet {} ...'.format(sample_id), end=' ', flush=True)
        start_action = datetime.now()
        append_df_to_excel(output_filename, 
                           df_val_params, 
                           sample_id, 
                           header=False)
        append_df_to_excel(output_filename, 
                           df_val, 
                           sample_id, 
                           index=False)
        stop_action = datetime.now()
        duration = stop_action - start_action
        print('Duration: {}'.format(duration.total_seconds()))
    else:
        print('A values sheet with Sample ID {} already exists in this workbook.'.format(sample_id))
        overwrite_sheet = query_yes_no('Do you want to overwrite this values sheet?')
        if overwrite_sheet:
            print('Overwriting values sheet {} ...'.format(sample_id), end=' ', flush=True)
            start_action = datetime.now()
            append_df_to_excel(output_filename, 
                               df_val_params, 
                               sample_id, 
                               startrow=0, 
                               header=False)
            append_df_to_excel(output_filename, 
                               df_val, 
                               sample_id, 
                               startrow=2, 
                               index=False)
            stop_action = datetime.now()
            duration = stop_action - start_action
            print('Duration: {}'.format(duration.total_seconds()))


def query_yes_no(question, default="yes"):
    """Ask a yes/no question via raw_input() and return their answer.

    "question" is a string that is presented to the user.
    "default" is the presumed answer if the user just hits <Enter>.
        It must be "yes" (the default), "no" or None (meaning
        an answer is required of the user).

    The "answer" return value is True for "yes" or False for "no".
    """
    valid = {"yes": True, "y": True, "ye": True,
             "no": False, "n": False}
    if default is None:
        prompt = " [y/n] "
    elif default == "yes":
        prompt = " [Y/n] "
    elif default == "no":
        prompt = " [y/N] "
    else:
        raise ValueError("invalid default answer: '%s'" % default)

    while True:
        sys.stdout.write(question + prompt)
        choice = input().lower()
        if default is not None and choice == '':
            return valid[default]
        elif choice in valid:
            return valid[choice]
        else:
            sys.stdout.write("Please respond with 'yes' or 'no' "
                             "(or 'y' or 'n').\n")


def append_df_to_excel(filename, 
                       df, 
                       sheet_name='Sheet1', 
                       startrow=None, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel`
                        [can be dictionary]

    Returns: None
    From: https://stackoverflow.com/questions/20219254/how-to-write-to-an-
    existing-excel-file-without-overwriting-data-using-pandas/47740262#47740262
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
            

if __name__ == '__main__':
    main(files)
