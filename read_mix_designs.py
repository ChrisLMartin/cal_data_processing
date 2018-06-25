# -*- coding: utf-8 -*-
"""
Created on Mon May 21 15:31:11 2018

@author: christopher.martin
"""

from openpyxl import load_workbook
import pandas as pd

in_file = "S:/Current Projects/R&D/ConcretePanTrials.xlsx"
#sheet = "20180516-01"

wb = load_workbook(in_file)
sheets = wb.sheetnames

# only import worksheets with mix designs on them
d = {}
for sheet in sheets:
    if "Chart" in sheet or "Results" in sheet or "ToBeCast" in sheet:
        continue
#    print(sheet)
    d[sheet] = pd.read_excel(in_file, sheet)

# select out 2 columns with the material names and quantities   
d2 = {}
for k, v in d.items():
    d2[k] = v.iloc[3:21, 1:3]

# manually sort funky one
d2["20171101-02"] = d["20171101-02"].iloc[3:21, :2]

# clean NaNs
for v in d2.values():
    v.dropna(inplace=True)

# make the material names the index, and the sample id the column name
for k, v in d2.items():
    v.columns = ["material", k]
    v.set_index("material", inplace=True)  

# convert to long data format 
d3 = {}
for k, v in d2.items():
    d3[k] = v.T

# test for duplicate columns
d4 = {}
for k, v in d3.items():
    if len(v.columns) != len(set(v.columns)):
        seen = {}
        dupes = []

        for x in v.columns:
            if x not in seen:
                seen[x] = 1
            else:
                if seen[x] == 1:
                    dupes.append(x)
                seen[x] += 1
        d4[k] = dupes

# concatenate all the series into single dataframe
result = pd.concat(d3, names=['dict key', 'index val']) # index as tuple (multi-index)
index = result.index.droplevel()
result.set_index(index, inplace=True)

writer = pd.ExcelWriter("mix_design_test.xlsx")
result.to_excel(writer, "Sheet1")
writer.save()
